"""
PokeBulb — Full TCGPlayer Product & Price Pull
===============================================
Fetches ALL Pokemon TCG sets from tcgcsv.com (Base Set through latest).
Saves:
  - all_products.csv          : full flat product+price data
  - all_products.xlsx         : same, formatted with era colour coding
  - sets_summary.csv          : one row per set, for reference
  - cache/<groupId>.json      : raw API cache per set (skip re-fetching)

Designed for:
  1. Site CSV import matching (HGSS onwards)
  2. PokeBulb API — productId lookup + live price fetching

Pricing formula:
  price_ZAR = ceil_to_nearest_0.50( market_price_USD * USD_ZAR_rate * 1.1 )

Usage:
  pip install pandas openpyxl requests
  python fetch_all_tcg_products.py

  # Fetch only specific eras:
  python fetch_all_tcg_products.py --era "Scarlet & Violet" "Mega Evolution"

  # Force re-fetch even if cached:
  python fetch_all_tcg_products.py --force
"""

import time, math, json, argparse, requests, pandas as pd, os
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUT_DIR   = Path(__file__).parent
CACHE_DIR = OUT_DIR / "tcg_cache"
CACHE_DIR.mkdir(exist_ok=True)

USER_AGENT   = "PokeBulbAPI/2.0"
CATEGORY_ID  = 3
MARKUP       = 1.1
SLEEP_MS     = 150
TIMEOUT      = 45
MAX_RETRIES  = 4

# ── All sets: (groupId, abbreviation, set_name, era) ─────────────────────────
ALL_SETS = [
    # Base / Fossil / Jungle era
    (604,  "BS",      "Base Set",                          "Base"),
    (1663, "BSS",     "Base Set (Shadowless)",             "Base"),
    (605,  "BS2",     "Base Set 2",                        "Base"),
    (635,  "JU",      "Jungle",                            "Base"),
    (630,  "FO",      "Fossil",                            "Base"),
    (1373, "TR",      "Team Rocket",                       "Base"),
    (1441, "G1",      "Gym Heroes",                        "Base"),
    (1440, "G2",      "Gym Challenge",                     "Base"),
    # Neo era
    (1396, "N1",      "Neo Genesis",                       "Neo"),
    (1434, "N2",      "Neo Discovery",                     "Neo"),
    (1389, "N3",      "Neo Revelation",                    "Neo"),
    (1444, "N4",      "Neo Destiny",                       "Neo"),
    # Legendary / e-Card era
    (1374, "LC",      "Legendary Collection",              "Legendary"),
    (1375, "EX",      "Expedition",                        "Legendary"),
    (1397, "AQ",      "Aquapolis",                         "Legendary"),
    (1372, "SK",      "Skyridge",                          "Legendary"),
    # EX era (Ruby/Sapphire)
    (1393, "RS",      "Ruby and Sapphire",                 "EX"),
    (1392, "SS",      "Sandstorm",                         "EX"),
    (1376, "DR",      "Dragon",                            "EX"),
    (1377, "MA",      "Team Magma vs Team Aqua",           "EX"),
    (1416, "HL",      "Hidden Legends",                    "EX"),
    (1419, "RG",      "FireRed & LeafGreen",               "EX"),
    (1428, "RR",      "Team Rocket Returns",               "EX"),
    (1404, "DX",      "Deoxys",                            "EX"),
    (1410, "EM",      "Emerald",                           "EX"),
    (1398, "UF",      "Unseen Forces",                     "EX"),
    (1429, "DS",      "Delta Species",                     "EX"),
    (1378, "LM",      "Legend Maker",                      "EX"),
    (1379, "HP",      "Holon Phantoms",                    "EX"),
    (1395, "CG",      "Crystal Guardians",                 "EX"),
    (1411, "DF",      "Dragon Frontiers",                  "EX"),
    (1383, "PK",      "Power Keepers",                     "EX"),
    # Diamond & Pearl era
    (1430, "DP",      "Diamond and Pearl",                 "DP"),
    (1368, "MT",      "Mysterious Treasures",              "DP"),
    (1405, "GE",      "Great Encounters",                  "DP"),
    (1390, "MD",      "Majestic Dawn",                     "DP"),
    (1417, "LA",      "Legends Awakened",                  "DP"),
    (1369, "SF",      "Stormfront",                        "DP"),
    # Platinum era
    (1406, "PL",      "Platinum",                          "Platinum"),
    (1367, "RR",      "Rising Rivals",                     "Platinum"),
    (1384, "SV",      "Supreme Victors",                   "Platinum"),
    (1391, "AR",      "Arceus",                            "Platinum"),
    # HG&SS era
    (1402, "HS",      "HeartGold SoulSilver",              "HG&SS"),
    (1399, "UL",      "Unleashed",                         "HG&SS"),
    (1403, "UD",      "Undaunted",                         "HG&SS"),
    (1381, "TM",      "Triumphant",                        "HG&SS"),
    (1415, "CoL",     "Call of Legends",                   "HG&SS"),
    # B&W era
    (1400, "BLW",     "Black and White",                   "B&W"),
    (1424, "EPO",     "Emerging Powers",                   "B&W"),
    (1385, "NVI",     "Noble Victories",                   "B&W"),
    (1412, "NXD",     "Next Destinies",                    "B&W"),
    (1386, "DEX",     "Dark Explorers",                    "B&W"),
    (1394, "DRX",     "Dragons Exalted",                   "B&W"),
    (1408, "BCR",     "Boundaries Crossed",                "B&W"),
    (1413, "PLS",     "Plasma Storm",                      "B&W"),
    (1382, "PLF",     "Plasma Freeze",                     "B&W"),
    (1370, "PLB",     "Plasma Blast",                      "B&W"),
    (1409, "LTR",     "Legendary Treasures",               "B&W"),
    # XY era
    (1387, "XY",      "XY Base Set",                       "XY"),
    (1464, "FLF",     "Flashfire",                         "XY"),
    (1481, "FFI",     "Furious Fists",                     "XY"),
    (1494, "PHF",     "Phantom Forces",                    "XY"),
    (1509, "PRC",     "Primal Clash",                      "XY"),
    (1525, "DCR",     "Double Crisis",                     "XY"),
    (1534, "ROS",     "Roaring Skies",                     "XY"),
    (1576, "AOR",     "Ancient Origins",                   "XY"),
    (1661, "BKT",     "BREAKthrough",                      "XY"),
    (1701, "BKP",     "BREAKpoint",                        "XY"),
    (1728, "GEN",     "Generations",                       "XY"),
    (1780, "FCO",     "Fates Collide",                     "XY"),
    (1815, "STS",     "Steam Siege",                       "XY"),
    (1842, "EVO",     "Evolutions",                        "XY"),
    # SM era
    (1863, "SM01",    "SM Base Set",                       "SM"),
    (1919, "SM02",    "Guardians Rising",                  "SM"),
    (1957, "SM03",    "Burning Shadows",                   "SM"),
    (2054, "SHL",     "Shining Legends",                   "SM"),
    (2071, "SM04",    "Crimson Invasion",                  "SM"),
    (2178, "SM05",    "Ultra Prism",                       "SM"),
    (2209, "SM06",    "Forbidden Light",                   "SM"),
    (2278, "CES",     "Celestial Storm",                   "SM"),
    (2295, "DRM",     "Dragon Majesty",                    "SM"),
    (2328, "SM8",     "Lost Thunder",                      "SM"),
    (2377, "SM9",     "Team Up",                           "SM"),
    (2420, "SM10",    "Unbroken Bonds",                    "SM"),
    (2464, "SM11",    "Unified Minds",                     "SM"),
    (2480, "HIF",     "Hidden Fates",                      "SM"),
    (2594, "HIFSV",   "Hidden Fates: Shiny Vault",         "SM"),
    (2534, "SM12",    "Cosmic Eclipse",                    "SM"),
    # SwSh era
    (2585, "SWSH01",  "Sword & Shield Base Set",           "SwSh"),
    (2626, "SWSH02",  "Rebel Clash",                       "SwSh"),
    (2675, "SWSH03",  "Darkness Ablaze",                   "SwSh"),
    (2685, "CHP",     "Champion's Path",                   "SwSh"),
    (2701, "SWSH04",  "Vivid Voltage",                     "SwSh"),
    (2754, "SHF",     "Shining Fates",                     "SwSh"),
    (2781, "SHFSV",   "Shining Fates: Shiny Vault",        "SwSh"),
    (2765, "SWSH05",  "Battle Styles",                     "SwSh"),
    (2807, "SWSH06",  "Chilling Reign",                    "SwSh"),
    (2848, "SWSH07",  "Evolving Skies",                    "SwSh"),
    (2867, "CLB",     "Celebrations",                      "SwSh"),
    (2931, "CCC",     "Celebrations: Classic Collection",  "SwSh"),
    (2906, "SWSH08",  "Fusion Strike",                     "SwSh"),
    (2948, "SWSH09",  "Brilliant Stars",                   "SwSh"),
    (3020, "BST",     "Brilliant Stars Trainer Gallery",   "SwSh"),
    (3040, "SWSH10",  "Astral Radiance",                   "SwSh"),
    (3068, "ASRTG",   "Astral Radiance Trainer Gallery",   "SwSh"),
    (3064, "PGO",     "Pokemon GO",                        "SwSh"),
    (3118, "SWSH11",  "Lost Origin",                       "SwSh"),
    (3172, "LORTG",   "Lost Origin Trainer Gallery",       "SwSh"),
    (3170, "SWSH12",  "Silver Tempest",                    "SwSh"),
    (17674,"ST",      "Silver Tempest Trainer Gallery",    "SwSh"),
    (17688,"CRZ",     "Crown Zenith",                      "SwSh"),
    (17689,"CRZGG",   "Crown Zenith: Galarian Gallery",    "SwSh"),
    # SV era
    (22873,"SVI",     "Scarlet & Violet Base Set",         "SV"),
    (22872,"SVP",     "Scarlet & Violet Promo Cards",      "SV"),
    (24382,"SVE",     "Scarlet & Violet Energies",         "SV"),
    (23120,"PAL",     "Paldea Evolved",                    "SV"),
    (23228,"OBF",     "Obsidian Flames",                   "SV"),
    (23237,"MEW",     "Scarlet & Violet 151",              "SV"),
    (23286,"PAR",     "Paradox Rift",                      "SV"),
    (23353,"PAF",     "Paldean Fates",                     "SV"),
    (23381,"TEF",     "Temporal Forces",                   "SV"),
    (23473,"TWM",     "Twilight Masquerade",               "SV"),
    (23529,"SFA",     "Shrouded Fable",                    "SV"),
    (23537,"SCR",     "Stellar Crown",                     "SV"),
    (23651,"SSP",     "Surging Sparks",                    "SV"),
    (23821,"PRE",     "Prismatic Evolutions",              "SV"),
    (24073,"JTG",     "Journey Together",                  "SV"),
    (24269,"DRI",     "Destined Rivals",                   "SV"),
    (24325,"BLK",     "Black Bolt",                        "SV"),
    (24326,"WHT",     "White Flare",                       "SV"),
    # Mega Evolution era
    (24380,"MEG",     "ME01: Mega Evolution",              "Mega Evolution"),
    (24461,"MEE",     "MEE: Mega Evolution Energies",      "Mega Evolution"),
    (24451,"MEP",     "ME: Mega Evolution Promo",          "Mega Evolution"),
    (24448,"PFL",     "ME02: Phantasmal Flames",           "Mega Evolution"),
    (24541,"ASC",     "ME: Ascended Heroes",               "Mega Evolution"),
    (24587,"POR",     "ME03: Perfect Order",               "Mega Evolution"),
    (24655,"CRI",     "ME04: Chaos Rising",                "Mega Evolution"),
    (24688,"ME05",    "ME05: Pitch Black",                 "Mega Evolution"),
]

ERA_COLORS = {
    "Base":           "5D6D7E",
    "Neo":            "717D7E",
    "Legendary":      "6E2EAB",
    "EX":             "1A5276",
    "DP":             "117A65",
    "Platinum":       "7D6608",
    "HG&SS":          "B7950B",
    "B&W":            "212F3D",
    "XY":             "7D3C98",
    "SM":             "1E8449",
    "SwSh":           "154360",
    "SV":             "922B21",
    "Mega Evolution":  "6E2FA1",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_usd_zar():
    """Fetch live USD/ZAR exchange rate."""
    try:
        r = requests.get(
            "https://api.frankfurter.app/latest?from=USD&to=ZAR",
            timeout=10, headers={"User-Agent": USER_AGENT}
        )
        rate = r.json()["rates"]["ZAR"]
        print(f"  💱 Live USD/ZAR rate: {rate:.4f}")
        return rate
    except Exception as e:
        print(f"  ⚠️  Could not fetch live rate ({e}). Using fallback R18.50")
        return 18.50

def ceil_50c(value):
    if value is None: return None
    return math.ceil(value * 2) / 2

def get_json(url, retries=MAX_RETRIES):
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT)
            r.raise_for_status()
            time.sleep(SLEEP_MS / 1000)
            return r.json()
        except Exception as e:
            if attempt < retries:
                wait = attempt * 6
                print(f"\n    ⚠️  Attempt {attempt} failed. Retrying in {wait}s...", end=" ", flush=True)
                time.sleep(wait)
            else:
                raise

def cache_path(group_id, kind):
    return CACHE_DIR / f"{group_id}_{kind}.json"

def fetch_with_cache(group_id, kind, url, force=False):
    cp = cache_path(group_id, kind)
    if cp.exists() and not force:
        return json.loads(cp.read_text(encoding="utf-8"))
    data = get_json(url)
    cp.write_text(json.dumps(data), encoding="utf-8")
    return data

def fetch_products(group_id, force=False):
    url  = f"https://tcgcsv.com/tcgplayer/{CATEGORY_ID}/{group_id}/products"
    data = fetch_with_cache(group_id, "products", url, force)
    out  = []
    for p in data.get("results", []):
        ext = {e["name"]: e["value"] for e in p.get("extendedData", [])}
        out.append({
            "productId":    p["productId"],
            "name":         p["name"],
            "cleanName":    p.get("cleanName", ""),
            "number":       ext.get("Number", ""),
            "rarity":       ext.get("Rarity", ""),
            "cardType":     ext.get("Card Type", ""),
            "hp":           ext.get("HP", ""),
            "stage":        ext.get("Stage", ""),
            "artist":       ext.get("Artist", ""),
            "isCard":       bool(ext.get("Number") or ext.get("Rarity")),
            "tcgplayer_url":p.get("url", ""),
        })
    return out

def fetch_prices(group_id, force=False):
    url  = f"https://tcgcsv.com/tcgplayer/{CATEGORY_ID}/{group_id}/prices"
    data = fetch_with_cache(group_id, "prices", url, force)
    prices = defaultdict(list)
    for p in data.get("results", []):
        prices[p["productId"]].append({
            "subTypeName":  p.get("subTypeName", "Normal"),
            "marketPrice":  p.get("marketPrice"),
            "lowPrice":     p.get("lowPrice"),
            "midPrice":     p.get("midPrice"),
            "highPrice":    p.get("highPrice"),
            "directLow":    p.get("directLowPrice"),
        })
    return prices

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--era",   nargs="+", help="Only fetch specific era(s)")
    parser.add_argument("--force", action="store_true", help="Ignore cache, re-fetch all")
    parser.add_argument("--rate",  type=float, help="Manual USD/ZAR rate (skip live fetch)")
    args = parser.parse_args()

    print("\n🃏 PokeBulb — Full TCG Product Pull")
    print("=" * 45)

    # Exchange rate
    usd_zar = args.rate if args.rate else get_usd_zar()
    print(f"  Pricing: USD × {MARKUP} × {usd_zar:.2f} ZAR → rounded to R0.50\n")

    # Filter sets
    sets_to_fetch = ALL_SETS
    if args.era:
        sets_to_fetch = [s for s in ALL_SETS if s[3] in args.era]
        print(f"  Filtering to eras: {args.era}")

    total = len(sets_to_fetch)
    all_rows = []
    set_summary = []
    failed = []

    print(f"  Fetching {total} sets...\n")

    for i, (gid, abbr, sname, era) in enumerate(sets_to_fetch, 1):
        print(f"  [{i:03d}/{total}] {era:<16} {abbr:<10} {sname}", end=" ... ", flush=True)

        # Check if fully cached
        cached = cache_path(gid, "products").exists() and cache_path(gid, "prices").exists()
        if cached and not args.force:
            print("♻️  cached", end=" ", flush=True)

        try:
            products = fetch_products(gid, args.force)
            prices   = fetch_prices(gid, args.force)

            cards      = [p for p in products if p["isCard"]]
            all_prods  = len(products)
            card_count = len(cards)
            price_rows = sum(len(v) for v in prices.values())

            for prod in products:
                pid      = prod["productId"]
                variants = prices.get(pid) or [{"subTypeName": "Normal", "marketPrice": None,
                                                 "lowPrice": None, "midPrice": None,
                                                 "highPrice": None, "directLow": None}]
                for pv in variants:
                    mkt = pv.get("marketPrice")
                    zar = None
                    if mkt is not None:
                        zar = ceil_50c(mkt * MARKUP * usd_zar)

                    all_rows.append({
                        # Identity
                        "era":           era,
                        "set_name":      sname,
                        "abbreviation":  abbr,
                        "group_id":      gid,
                        "productId":     pid,
                        # Card info
                        "name":          prod["name"],
                        "cleanName":     prod["cleanName"],
                        "number":        prod["number"],
                        "rarity":        prod["rarity"],
                        "cardType":      prod["cardType"],
                        "hp":            prod["hp"],
                        "stage":         prod["stage"],
                        "artist":        prod["artist"],
                        "isCard":        prod["isCard"],
                        # Pricing
                        "subTypeName":   pv["subTypeName"],
                        "market_usd":    mkt,
                        "low_usd":       pv.get("lowPrice"),
                        "mid_usd":       pv.get("midPrice"),
                        "high_usd":      pv.get("highPrice"),
                        "direct_low_usd":pv.get("directLow"),
                        "usd_zar_rate":  round(usd_zar, 4),
                        "pokebulb_zar":  zar,
                        # URL
                        "tcgplayer_url": prod["tcgplayer_url"],
                    })

            set_summary.append({
                "era": era, "set_name": sname, "abbreviation": abbr,
                "group_id": gid, "total_products": all_prods,
                "cards_only": card_count, "price_rows": price_rows,
                "cached": cached and not args.force,
            })

            print(f"✅ {card_count} cards, {price_rows} price rows")

        except Exception as e:
            print(f"❌ ERROR: {e}")
            failed.append((gid, abbr, sname, str(e)))
            set_summary.append({
                "era": era, "set_name": sname, "abbreviation": abbr,
                "group_id": gid, "total_products": 0,
                "cards_only": 0, "price_rows": 0, "cached": False,
            })

    # ── Save outputs ──────────────────────────────────────────────────────────
    df = pd.DataFrame(all_rows)
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    csv_path = OUT_DIR / f"all_tcg_products_{ts}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"\n✅ CSV saved  → {csv_path.name}")

    sum_df = pd.DataFrame(set_summary)
    sum_path = OUT_DIR / "sets_summary.csv"
    sum_df.to_csv(sum_path, index=False, encoding="utf-8-sig")
    print(f"✅ Summary   → {sum_path.name}")

    # ── Excel ─────────────────────────────────────────────────────────────────
    build_excel(df, OUT_DIR / f"all_tcg_products_{ts}.xlsx")
    print(f"✅ Excel saved → all_tcg_products_{ts}.xlsx")

    # ── Final report ─────────────────────────────────────────────────────────
    print(f"\n{'='*45}")
    print(f"  Total rows:        {len(df):,}")
    print(f"  Unique productIds: {df['productId'].nunique():,}")
    print(f"  Sets completed:    {len(set_summary) - len(failed)}/{total}")
    print(f"  USD/ZAR used:      {usd_zar:.4f}")

    if failed:
        print(f"\n  ❌ Failed sets ({len(failed)}) — run again to retry:")
        for gid, abbr, sname, err in failed:
            print(f"     {abbr:<10} {sname}  ({err[:60]})")
    else:
        print(f"\n  🎉 All sets fetched successfully!")

    print(f"\n  Output folder: {OUT_DIR}\n")


def build_excel(df, path):
    wb  = Workbook()
    thin = Side(style="thin", color="DDDDDD")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(cell, color="1F4E79"):
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr

    def dat(cell, bold=False, fg="000000", bg=None, mono=False):
        cell.font      = Font(name="Courier New" if mono else "Arial", size=9,
                              bold=bold, color=fg)
        cell.alignment = Alignment(vertical="center")
        cell.border    = bdr
        if bg:
            cell.fill  = PatternFill("solid", fgColor=bg)

    # ── Sheet 1: All products ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Products"

    cols = [
        ("Era",          "era"),
        ("Set",          "set_name"),
        ("Abbrev",       "abbreviation"),
        ("Group ID",     "group_id"),
        ("Product ID",   "productId"),
        ("Name",         "name"),
        ("Number",       "number"),
        ("Rarity",       "rarity"),
        ("Type",         "cardType"),
        ("Stage",        "stage"),
        ("Variant",      "subTypeName"),
        ("Market USD",   "market_usd"),
        ("Low USD",      "low_usd"),
        ("PokeBulb ZAR", "pokebulb_zar"),
        ("USD/ZAR Rate", "usd_zar_rate"),
        ("Is Card",      "isCard"),
    ]
    labels = [c[0] for c in cols]
    keys   = [c[1] for c in cols]

    ws.append(labels)
    for ci in range(1, len(labels)+1):
        hdr(ws.cell(1, ci))

    for ri, row in enumerate(df[keys].itertuples(index=False), 2):
        era_c = ERA_COLORS.get(row[0], "444444")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, value=val)
            if ci == 1:
                dat(cell, bold=True, bg=era_c, fg="FFFFFF")
            elif ci in (4, 5):
                dat(cell, bold=True, fg="1A5276", mono=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 12 and val is not None:
                cell.number_format = '"$"#,##0.00'
                dat(cell)
            elif ci == 14 and val is not None:
                cell.number_format = '"R"#,##0.00'
                dat(cell, bold=True, fg="1A6B3C")
            else:
                dat(cell)

    widths = [16, 32, 9, 10, 12, 36, 10, 22, 12, 12, 18, 12, 10, 14, 12, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(labels))}1"

    # ── Sheet 2: Era summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Era Summary")
    era_sum = df.groupby("era").agg(
        sets=("abbreviation", "nunique"),
        products=("productId", "nunique"),
        rows=("productId", "count"),
        priced=("market_usd", lambda x: x.notna().sum()),
    ).reset_index()
    ws2.append(["Era", "Sets", "Unique Products", "Total Rows", "Rows with Price"])
    for ci in range(1, 6): hdr(ws2.cell(1, ci))
    for ri, r in enumerate(era_sum.itertuples(index=False), 2):
        ec = ERA_COLORS.get(r.era, "444444")
        for ci, val in enumerate(r, 1):
            cell = ws2.cell(ri, ci, value=val)
            if ci == 1:
                dat(cell, bold=True, bg=ec, fg="FFFFFF")
            else:
                dat(cell)
    for ci, w in enumerate([20,8,16,12,16], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)


if __name__ == "__main__":
    main()
