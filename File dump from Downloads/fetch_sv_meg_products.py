"""
pokebulb — Full Product ID Pull: Scarlet & Violet + Mega Evolution
==================================================================
Run this locally (tcgcsv.com is not accessible from Claude's sandbox).

    pip install pandas openpyxl requests
    python fetch_sv_meg_products.py

Outputs:
  sv_meg_product_ids.xlsx   — full product list with IDs, variants, prices
  sv_meg_product_ids.csv    — flat CSV version for VLOOKUP in your store file
"""

import time, math, requests, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Sets to fetch ─────────────────────────────────────────────────────────────
# groupId → (abbreviation, set_name, era)
TARGET_SETS = {
    # ── Scarlet & Violet ──────────────────────────────────────────────────────
    24325: ("BLK",    "SV: Black Bolt",                  "Scarlet & Violet"),
    24326: ("WHT",    "SV: White Flare",                 "Scarlet & Violet"),
    24269: ("DRI",    "SV10: Destined Rivals",           "Scarlet & Violet"),
    24073: ("JTG",    "SV09: Journey Together",          "Scarlet & Violet"),
    23821: ("PRE",    "SV: Prismatic Evolutions",        "Scarlet & Violet"),
    23651: ("SSP",    "SV08: Surging Sparks",            "Scarlet & Violet"),
    23537: ("SCR",    "SV07: Stellar Crown",             "Scarlet & Violet"),
    23529: ("SFA",    "SV: Shrouded Fable",              "Scarlet & Violet"),
    23473: ("TWM",    "SV06: Twilight Masquerade",       "Scarlet & Violet"),
    23381: ("TEF",    "SV05: Temporal Forces",           "Scarlet & Violet"),
    23353: ("PAF",    "SV: Paldean Fates",               "Scarlet & Violet"),
    23286: ("PAR",    "SV04: Paradox Rift",              "Scarlet & Violet"),
    23237: ("MEW",    "SV: Scarlet & Violet 151",        "Scarlet & Violet"),
    23228: ("OBF",    "SV03: Obsidian Flames",           "Scarlet & Violet"),
    23120: ("PAL",    "SV02: Paldea Evolved",            "Scarlet & Violet"),
    22873: ("SVI",    "SV01: Scarlet & Violet Base Set", "Scarlet & Violet"),
    22872: ("SVP",    "SV: Scarlet & Violet Promo Cards","Scarlet & Violet"),
    24382: ("SVE",    "SVE: Scarlet & Violet Energies",  "Scarlet & Violet"),
    # ── Mega Evolution ────────────────────────────────────────────────────────
    24688: ("ME05",   "ME05: Pitch Black",               "Mega Evolution"),
    24655: ("CRI",    "ME04: Chaos Rising",              "Mega Evolution"),
    24587: ("POR",    "ME03: Perfect Order",             "Mega Evolution"),
    24541: ("ASC",    "ME: Ascended Heroes",             "Mega Evolution"),
    24448: ("PFL",    "ME02: Phantasmal Flames",         "Mega Evolution"),
    24451: ("MEP",    "ME: Mega Evolution Promo",        "Mega Evolution"),
    24380: ("MEG",    "ME01: Mega Evolution",            "Mega Evolution"),
    24461: ("MEE",    "MEE: Mega Evolution Energies",    "Mega Evolution"),
}

CATEGORY_ID = 3
USER_AGENT  = "PokebulbProductPull/1.0"
SLEEP_MS    = 150   # polite rate limiting

ERA_COLORS = {
    "Scarlet & Violet": "C0392B",
    "Mega Evolution":   "6E2FA1",
}

def get_json(url):
    r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=20)
    r.raise_for_status()
    time.sleep(SLEEP_MS / 1000)
    return r.json()

def fetch_products(group_id):
    url  = f"https://tcgcsv.com/tcgplayer/{CATEGORY_ID}/{group_id}/products"
    data = get_json(url)
    out  = []
    for p in data.get("results", []):
        ext = {e["name"]: e["value"] for e in p.get("extendedData", [])}
        # Only include actual cards (have Number or Rarity)
        if not ext.get("Number") and not ext.get("Rarity"):
            continue
        out.append({
            "productId":   p["productId"],
            "name":        p["name"],
            "cleanName":   p.get("cleanName", ""),
            "number":      ext.get("Number", ""),
            "rarity":      ext.get("Rarity", ""),
            "cardType":    ext.get("Card Type", ""),
            "hp":          ext.get("HP", ""),
            "stage":       ext.get("Stage", ""),
            "tcgplayer_url": p.get("url", ""),
        })
    return out

def fetch_prices(group_id):
    url  = f"https://tcgcsv.com/tcgplayer/{CATEGORY_ID}/{group_id}/prices"
    data = get_json(url)
    # Return dict: productId → list of {subTypeName, marketPrice, lowPrice, midPrice}
    from collections import defaultdict
    prices = defaultdict(list)
    for p in data.get("results", []):
        prices[p["productId"]].append({
            "subTypeName":  p.get("subTypeName", "Normal"),
            "marketPrice":  p.get("marketPrice"),
            "lowPrice":     p.get("lowPrice"),
            "midPrice":     p.get("midPrice"),
            "highPrice":    p.get("highPrice"),
        })
    return prices

def ceil_50c(v):
    if v is None: return None
    return math.ceil(v * 2) / 2

# ── Main fetch loop ───────────────────────────────────────────────────────────
all_rows = []
total = len(TARGET_SETS)

print(f"Fetching {total} sets...\n")
for i, (gid, (abbr, sname, era)) in enumerate(TARGET_SETS.items(), 1):
    print(f"[{i:02d}/{total}] {sname} ({abbr})...", end=" ", flush=True)
    try:
        products = fetch_products(gid)
        prices   = fetch_prices(gid)
        for prod in products:
            pid = prod["productId"]
            pvariants = prices.get(pid, [{"subTypeName": "Normal", "marketPrice": None, "lowPrice": None, "midPrice": None, "highPrice": None}])
            for pv in pvariants:
                mkt = pv.get("marketPrice")
                all_rows.append({
                    "era":          era,
                    "set_name":     sname,
                    "abbreviation": abbr,
                    "group_id":     gid,
                    "productId":    pid,
                    "name":         prod["name"],
                    "cleanName":    prod["cleanName"],
                    "number":       prod["number"],
                    "rarity":       prod["rarity"],
                    "cardType":     prod["cardType"],
                    "hp":           prod["hp"],
                    "stage":        prod["stage"],
                    "subTypeName":  pv["subTypeName"],
                    "marketPrice":  mkt,
                    "lowPrice":     pv.get("lowPrice"),
                    "midPrice":     pv.get("midPrice"),
                    "pokebulb_price": ceil_50c(mkt * 1.1) if mkt else None,
                    "tcgplayer_url": prod["tcgplayer_url"],
                })
        print(f"{len(products)} cards, {sum(len(v) for v in prices.values())} price rows")
    except Exception as e:
        print(f"ERROR: {e}")

df = pd.DataFrame(all_rows)

# Sort: era, then set publish order (use groupId order from TARGET_SETS), then card number
set_order = {gid: i for i, gid in enumerate(TARGET_SETS.keys())}
df["_set_order"] = df["group_id"].map(set_order)

def num_sort_key(n):
    try: return int(str(n).split("/")[0])
    except: return 9999

df["_num_sort"] = df["number"].apply(num_sort_key)
df = df.sort_values(["era", "_set_order", "_num_sort", "subTypeName"]).drop(columns=["_set_order","_num_sort"])

# ── Save CSV ──────────────────────────────────────────────────────────────────
csv_out = "sv_meg_product_ids.csv"
df.drop(columns=["tcgplayer_url"]).to_csv(csv_out, index=False, encoding="utf-8-sig")
print(f"\n✅ CSV saved: {csv_out}")

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = Workbook()

thin = Side(style="thin", color="CCCCCC")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_style(cell, hex_color="1F4E79"):
    cell.fill = PatternFill("solid", fgColor=hex_color)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr

def data_style(cell, bold=False, bg=None, fg="000000"):
    cell.font = Font(name="Arial", size=9, bold=bold, color=fg)
    cell.alignment = Alignment(vertical="center")
    cell.border = bdr
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

# ── Sheet 1: Full Product List ────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "All Products"

display_cols = [
    ("Era",            "era"),
    ("Set Name",       "set_name"),
    ("Abbrev.",        "abbreviation"),
    ("Group ID",       "group_id"),
    ("Product ID",     "productId"),
    ("Card Name",      "name"),
    ("Number",         "number"),
    ("Rarity",         "rarity"),
    ("Card Type",      "cardType"),
    ("Stage",          "stage"),
    ("Variant/SubType","subTypeName"),
    ("Market $USD",    "marketPrice"),
    ("Low $USD",       "lowPrice"),
    ("Mid $USD",       "midPrice"),
    ("Pokebulb R (1.1x↑0.50)", "pokebulb_price"),
]

col_labels = [c[0] for c in display_cols]
col_keys   = [c[1] for c in display_cols]

ws1.append(col_labels)
for ci, label in enumerate(col_labels, 1):
    hdr_style(ws1.cell(row=1, column=ci))

for ri, row in enumerate(df[col_keys].itertuples(index=False), 2):
    era_val = row[0]
    era_color = ERA_COLORS.get(era_val, "444444")
    for ci, val in enumerate(row, 1):
        cell = ws1.cell(row=ri, column=ci, value=val)
        if ci == 1:  # Era column coloured
            data_style(cell, bold=True, bg=era_color, fg="FFFFFF")
        elif ci in (4, 5):  # Group ID / Product ID
            data_style(cell, bold=True, fg="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 12 and val is not None:  # Market price
            cell.number_format = '"$"#,##0.00'
            data_style(cell)
        elif ci == 15 and val is not None:  # Pokebulb price
            cell.number_format = '"R"#,##0.00'
            data_style(cell, bold=True, fg="1A6B3C")
        else:
            data_style(cell)

col_widths_1 = [18, 34, 9, 10, 12, 32, 10, 18, 14, 12, 16, 12, 10, 10, 22]
for ci, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

ws1.row_dimensions[1].height = 30
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(col_labels))}1"

# ── Sheet 2: Per-set summary ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Set Summary")
s_headers = ["Era","Set Name","Abbrev.","Group ID","Card Count","Variant Rows","Subtypes Present","Sets with Prices"]
ws2.append(s_headers)
for ci in range(1, len(s_headers)+1):
    hdr_style(ws2.cell(row=1, column=ci))

summary = df.groupby(["era","set_name","abbreviation","group_id"], sort=False).agg(
    card_count=("productId","nunique"),
    variant_rows=("productId","count"),
    subtypes=("subTypeName", lambda x: " | ".join(sorted(x.unique()))),
    has_prices=("marketPrice", lambda x: "Yes" if x.notna().any() else "No"),
).reset_index()

for ri, row in enumerate(summary.itertuples(index=False), 2):
    vals = [row.era, row.set_name, row.abbreviation, row.group_id,
            row.card_count, row.variant_rows, row.subtypes, row.has_prices]
    era_color = ERA_COLORS.get(row.era, "444444")
    for ci, val in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        if ci == 1:
            data_style(cell, bold=True, bg=era_color, fg="FFFFFF")
        elif ci == 4:
            data_style(cell, bold=True, fg="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            data_style(cell)

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 36
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 50
ws2.column_dimensions["H"].width = 14
ws2.row_dimensions[1].height = 28
ws2.freeze_panes = "A2"

# ── Sheet 3: Rarity / Variant breakdown ──────────────────────────────────────
ws3 = wb.create_sheet("Variants by Set")
v_headers = ["Era","Set Name","Abbrev.","Rarity","SubType / Variant","Count"]
ws3.append(v_headers)
for ci in range(1, len(v_headers)+1):
    hdr_style(ws3.cell(row=1, column=ci))

var_df = df.groupby(["era","set_name","abbreviation","rarity","subTypeName"], sort=False).size().reset_index(name="count")
for ri, row in enumerate(var_df.itertuples(index=False), 2):
    vals = [row.era, row.set_name, row.abbreviation, row.rarity, row.subTypeName, row.count]
    era_color = ERA_COLORS.get(row.era, "444444")
    for ci, val in enumerate(vals, 1):
        cell = ws3.cell(row=ri, column=ci, value=val)
        if ci == 1:
            data_style(cell, bold=True, bg=era_color, fg="FFFFFF")
        else:
            data_style(cell)

for ci, w in enumerate([18,36,10,20,20,8], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[1].height = 28
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:{get_column_letter(len(v_headers))}1"

xl_out = "sv_meg_product_ids.xlsx"
wb.save(xl_out)

print(f"✅ Excel saved: {xl_out}")
print(f"\nTotal rows: {len(df)}")
print(f"Unique products: {df['productId'].nunique()}")
print(f"Unique sets: {df['set_name'].nunique()}")
print("\nVariants found across all sets:")
print(df["subTypeName"].value_counts().to_string())
