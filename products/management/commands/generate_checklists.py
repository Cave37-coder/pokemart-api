import os
import requests
import time

from django.core.management.base import BaseCommand
from django.conf import settings

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

TCG_API_BASE = "https://api.pokemontcg.io/v2"
TCG_API_KEY  = os.environ.get("POKEMON_TCG_API_KEY", "")

RARITY_SYMBOL = {
    "Common": "●", "Uncommon": "◆", "Rare": "★",
    "Rare Holo": "★H", "Rare Holo EX": "★H", "Rare Holo GX": "★H",
    "Rare Holo V": "★H", "Rare Holo VMAX": "★H", "Rare Holo VSTAR": "★H",
    "Illustration Rare": "IR", "Special Illustration Rare": "SIR",
    "Ultra Rare": "UR", "Secret Rare": "SR", "Hyper Rare": "HR",
    "Promo": "PR", "Amazing Rare": "AR", "Double Rare": "RR",
    "Shiny Rare": "SHR", "ACE SPEC Rare": "ACE",
}


def get_headers():
    h = {"User-Agent": "PokeBulkSA-checklist/1.0"}
    if TCG_API_KEY:
        h["X-Api-Key"] = TCG_API_KEY
    return h


def fetch_set_info(set_id):
    r = requests.get(f"{TCG_API_BASE}/sets/{set_id}", headers=get_headers(), timeout=15)
    r.raise_for_status()
    return r.json()["data"]


def fetch_all_cards(set_id):
    cards, page = [], 1
    while True:
        url = (f"{TCG_API_BASE}/cards?q=set.id:{set_id}"
               f"&page={page}&pageSize=250&orderBy=number"
               f"&select=number,name,rarity,types,supertype")
        r = requests.get(url, headers=get_headers(), timeout=15)
        r.raise_for_status()
        data = r.json()
        cards.extend(data["data"])
        if len(cards) >= data["totalCount"]:
            break
        page += 1
        time.sleep(0.3)
    return cards


def sort_cards(cards):
    def key(c):
        num = c.get("number", "0")
        try:
            return (int(num), "")
        except ValueError:
            digits = "".join(filter(str.isdigit, num))
            return (int(digits) if digits else 9999, num)
    return sorted(cards, key=key)


def output_dir():
    path = os.path.join(settings.MEDIA_ROOT, "checklists")
    os.makedirs(path, exist_ok=True)
    return path


def generate_pdf(set_id, set_name, set_series, cards, out_path):
    doc = SimpleDocTemplate(out_path, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm)

    mk = lambda name, **kw: ParagraphStyle(name, **kw)
    s_title = mk("T", fontSize=22, fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=2)
    s_sub   = mk("S", fontSize=11, fontName="Helvetica", alignment=TA_CENTER, spaceAfter=6)
    s_hint  = mk("H", fontSize=9, fontName="Helvetica-Oblique", alignment=TA_CENTER, spaceAfter=10)
    s_cell  = mk("C", fontSize=7.5, fontName="Helvetica", leading=10)
    s_sym   = mk("SY", fontSize=7.5, fontName="Helvetica", alignment=TA_CENTER)
    s_hdr   = mk("HD", fontSize=7, fontName="Helvetica-Bold")
    s_hdrc  = mk("HDC", fontSize=7, fontName="Helvetica-Bold", alignment=TA_CENTER)
    s_leg   = mk("L", fontSize=8, fontName="Helvetica", alignment=TA_CENTER, spaceAfter=4)
    s_foot  = mk("F", fontSize=7, fontName="Helvetica", textColor=colors.grey, alignment=TA_CENTER)

    story = [
        Paragraph(set_name.upper(), s_title),
        Paragraph(f"{set_series} - Complete Card Checklist", s_sub),
        Paragraph("Use the check boxes to keep track of your Pokemon TCG cards!", s_hint),
    ]

    rows_per_col = -(-len(cards) // 3)
    col_data = [[], [], []]
    for i, card in enumerate(cards):
        ci = min(i // rows_per_col, 2)
        col_data[ci].append([
            f"[] {card.get('number',''):>4}",
            card.get("name", ""),
            RARITY_SYMBOL.get(card.get("rarity", ""), ""),
        ])

    max_rows = max(len(c) for c in col_data)
    for col in col_data:
        while len(col) < max_rows:
            col.append(["", "", ""])

    table_rows = [[
        Paragraph("#", s_hdr), Paragraph("Name", s_hdr), Paragraph("Rarity", s_hdrc), Paragraph("", s_hdr),
        Paragraph("#", s_hdr), Paragraph("Name", s_hdr), Paragraph("Rarity", s_hdrc), Paragraph("", s_hdr),
        Paragraph("#", s_hdr), Paragraph("Name", s_hdr), Paragraph("Rarity", s_hdrc),
    ]]

    for i in range(max_rows):
        row = []
        for ci in range(3):
            num, name, sym = col_data[ci][i]
            row += [Paragraph(num, s_cell), Paragraph(name, s_cell), Paragraph(sym, s_sym)]
            if ci < 2:
                row.append(Paragraph("", s_cell))
        table_rows.append(row)

    page_w = A4[0] - 30*mm
    cw_raw = [14, 48, 12, 5, 14, 48, 12, 5, 14, 48, 12]
    scale  = page_w / (sum(cw_raw) * mm)
    cw     = [c * mm * scale for c in cw_raw]

    tbl = Table(table_rows, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME",       (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0,0), (-1,-1), 7.5),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.Color(0.96,0.96,0.96)]),
        ("LINEBELOW",      (0,0), (-1,0),  0.5, colors.black),
        ("TOPPADDING",     (0,0), (-1,-1), 1.5),
        ("BOTTOMPADDING",  (0,0), (-1,-1), 1.5),
        ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
    ]))

    story += [
        tbl,
        Spacer(1, 6*mm),
        Paragraph(
            "Symbol Legend: circle=Common  diamond=Uncommon  star=Rare  "
            "starH=Rare Holo  IR=Illustration Rare  UR=Ultra Rare  SR=Secret Rare  PR=Promo",
            s_leg,
        ),
        Spacer(1, 3*mm),
        Paragraph(f"Generated by PokeBulk SA - pokebulk.co.za - {len(cards)} cards", s_foot),
    ]

    doc.build(story)
    print(f"  PDF saved: {out_path}")


def generate_excel(set_id, set_name, set_series, cards, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = set_name[:31]

    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    for rn, (text, sz, bold, italic, col) in enumerate([
        (f"{set_name} - Complete Card Checklist", 14, True,  False, "000000"),
        (f"{set_series} - {len(cards)} cards - pokebulk.co.za", 9, False, True, "888888"),
        ('Type anything in "Have it" column - row turns green automatically.', 9, False, True, "555555"),
    ], 1):
        ws.merge_cells(f"A{rn}:G{rn}")
        cell = ws.cell(row=rn, column=1, value=text)
        cell.font      = Font(name="Arial", bold=bold, italic=italic, size=sz, color=col)
        cell.alignment = center
        ws.row_dimensions[rn].height = 20 if rn == 1 else 14

    headers    = ["#", "Name", "Supertype", "Type(s)", "Rarity", "Symbol", "Have it"]
    col_widths = [7,    28,     12,          16,        22,       8,        10]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color="1A1A2E")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 20
    ws.freeze_panes = "A5"

    alt_fill = PatternFill("solid", start_color="F4F4F8")
    for ri, card in enumerate(cards, 5):
        rarity = card.get("rarity", "")
        vals = [
            card.get("number", ""),
            card.get("name", ""),
            card.get("supertype", ""),
            ", ".join(card.get("types", []) or []),
            rarity,
            RARITY_SYMBOL.get(rarity, ""),
            "",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=11 if ci == 7 else 9)
            cell.border    = border
            cell.alignment = center if ci in (1, 6, 7) else left
            if ri % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[ri].height = 16

    last_row = 4 + len(cards)
    ws.conditional_formatting.add(
        f"A5:G{last_row}",
        FormulaRule(
            formula=['$G5<>""'],
            fill=PatternFill("solid", start_color="C3E6CB"),
            font=Font(name="Arial", size=9, color="155724"),
        )
    )

    notes = wb.create_sheet("How to use")
    notes["A1"] = "How to use this checklist"
    notes["A1"].font = Font(name="Arial", bold=True, size=13)
    for i, line in enumerate([
        "",
        "1. Type anything in the Have it column (G) when you own a card.",
        "2. The row automatically turns green.",
        "3. Filter Have it for blanks to find cards you still need.",
        "",
        "Set tiers:",
        "  Simple set   = Commons and Holos",
        "  Base set     = All numbered cards including Reverse Holos",
        "  Master set   = Base set + Illustration Rares",
        "  Grand Master = Master set + Promos",
        "",
        "Visit pokebulk.co.za to shop for your missing cards!",
    ], 2):
        notes.cell(row=i, column=1, value=line).font = Font(name="Arial", size=10)
    notes.column_dimensions["A"].width = 70

    wb.save(out_path)
    print(f"  Excel saved: {out_path}")


class Command(BaseCommand):
    help = "Generate PDF + Excel checklists per set using Pokemon TCG API"

    def add_arguments(self, parser):
        parser.add_argument("--set", dest="set_id", type=str,
                            help="Single set ID e.g. swsh1")
        parser.add_argument("--skip-existing", action="store_true", default=True)

    def handle(self, *args, **options):
        from products.models import CardSet

        sid_filter    = options.get("set_id")
        skip_existing = options.get("skip_existing")
        out           = output_dir()

        qs = (CardSet.objects.filter(code=sid_filter)
              if sid_filter else CardSet.objects.all().order_by("release_date"))

        if not qs.exists():
            self.stdout.write(self.style.ERROR("No sets found."))
            return

        self.stdout.write(f"Processing {qs.count()} set(s)...")

        for card_set in qs:
            sid       = card_set.set_id
            pdf_path  = os.path.join(out, f"{sid}_checklist.pdf")
            xlsx_path = os.path.join(out, f"{sid}_checklist.xlsx")

            if skip_existing and os.path.exists(pdf_path) and os.path.exists(xlsx_path):
                self.stdout.write(f"  Skipping {sid} (files exist)")
                continue

            self.stdout.write(f"\n-> {sid}...")
            try:
                info   = fetch_set_info(sid)
                name   = info.get("name", card_set.name)
                series = info.get("series", "")
                cards  = sort_cards(fetch_all_cards(sid))
                self.stdout.write(f"  {len(cards)} cards fetched")

                generate_pdf(sid, name, series, cards, pdf_path)
                generate_excel(sid, name, series, cards, xlsx_path)

                card_set.checklist_pdf  = f"checklists/{sid}_checklist.pdf"
                card_set.checklist_xlsx = f"checklists/{sid}_checklist.xlsx"
                card_set.save(update_fields=["checklist_pdf", "checklist_xlsx"])

                time.sleep(0.5)

            except requests.HTTPError as e:
                self.stdout.write(self.style.ERROR(f"  HTTP error: {e}"))
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"  Failed: {e}"))

        self.stdout.write(self.style.SUCCESS("\nDone! Files in /media/checklists/"))
