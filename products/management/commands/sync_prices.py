import math, os, time, unicodedata, requests
from decimal import Decimal
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from products.models import CardSet, PokemonProduct

TCGCSV_BASE = "https://tcgcsv.com/tcgplayer"
POKEMON_CATEGORY = 3
EXCHANGE_RATE_API = "https://api.exchangerate-api.com/v4/latest/USD"
MARKUP = Decimal("1.1")
HEADERS = {"User-Agent": "PokeBulkSA/1.0"}
ALL_PRICE_FIELDS = ["price_normal", "price_holo", "price_reverse_holo", "price_first_edition", "price_pokeball", "price_masterball", "price_friendball", "price_loveball", "price_quickball", "price_duskball"]

VARIANT_TO_FIELD = {
    "Normal": "price_normal", "Unlimited": "price_normal", "Unlimited Normal": "price_normal",
    "Holofoil": "price_holo", "Unlimited Holofoil": "price_holo",
    "Reverse Holofoil": "price_reverse_holo",
    "Poke Ball": "price_pokeball", "Poke Ball Holofoil": "price_pokeball",
    "Master Ball": "price_masterball", "Master Ball Holofoil": "price_masterball",
    "Friend Ball": "price_friendball", "Friend Ball Holofoil": "price_friendball",
    "Love Ball": "price_loveball", "Love Ball Holofoil": "price_loveball",
    "Quick Ball": "price_quickball", "Quick Ball Holofoil": "price_quickball",
    "Dusk Ball": "price_duskball", "Dusk Ball Holofoil": "price_duskball",
    "1st Edition": "price_first_edition", "1st Edition Normal": "price_first_edition",
    "1st Edition Holofoil": "price_first_edition",
}

NAME_FIXES = {
    "imposter professor oak": "impostor professor oak",
    "nidoran m": "nidoran \u2642",
}

# xlsx set code -> DB set code
XLSX_TO_DB = {
    "CoL":    "CL",
    "SVI":    "SV1",
    "SVP":    "PR-SV",
    "SM01":   "SUM",
    "SM02":   "GRI",
    "SM03":   "BUS",
    "SM04":   "CIN",
    "SM05":   "UPR",
    "SM06":   "FLI",
    "SM8":    "LOT",
    "SM9":    "TRR",
    "SM10":   "UPR",
    "SM11":   "UPR",
    "SM12":   "UPR",
    "SHL":    "SLG",
    "SWSH01": "SV1",
    "SWSH02": "SV1",
    "SWSH03": "SV1",
    "SWSH04": "SV1",
    "SWSH05": "SV1",
    "SWSH06": "SV1",
    "SWSH07": "SV1",
    "SWSH08": "SV1",
    "SWSH09": "SV1",
    "SWSH10": "SV1",
    "SWSH11": "SV1",
    "SWSH12": "SV1",
}

def round_up_50c(zar):
    return Decimal(math.ceil(zar * 2)) / 2

def to_zar(usd, rate):
    return round_up_50c(usd * rate * MARKUP)

def norm_num(raw):
    return str(raw or "").split("/")[0].lstrip("0") or "0"

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

VARIANT_SUFFIXES = [
    " (holo)", " (reverse holo)", " (normal)", " (holofoil)",
    " (1st edition)", " (unlimited)", " (reverse holofoil)",
    " (poke ball reverse holo)", " (master ball reverse holo)",
    " (friend ball reverse holo)", " (love ball reverse holo)",
    " (quick ball reverse holo)", " (dusk ball reverse holo)",
    " (pokeball reverse holo)", " (masterball reverse holo)",
]

def norm_name(name):
    n = strip_accents(str(name).strip().lower())
    for suffix in VARIANT_SUFFIXES:
        if n.endswith(suffix):
            n = n[:-len(suffix)].strip()
            break
    return NAME_FIXES.get(n, n)

def name_key(name, number):
    return f"{norm_name(name)}|{norm_num(number)}"

def variant_to_field(variant):
    f = VARIANT_TO_FIELD.get(variant)
    if f: return f
    vl = (variant or "").lower()
    if "poke ball" in vl: return "price_pokeball"
    if "master ball" in vl: return "price_masterball"
    if "friend ball" in vl: return "price_friendball"
    if "love ball" in vl: return "price_loveball"
    if "quick ball" in vl: return "price_quickball"
    if "dusk ball" in vl: return "price_duskball"
    if "reverse" in vl: return "price_reverse_holo"
    if "1st" in vl or "first" in vl: return "price_first_edition"
    if "holo" in vl: return "price_holo"
    return "price_normal"

def fetch_rate():
    r = requests.get(EXCHANGE_RATE_API, timeout=10, headers=HEADERS)
    r.raise_for_status()
    return Decimal(str(r.json()["rates"]["ZAR"]))

def fetch_groups():
    r = requests.get(f"{TCGCSV_BASE}/{POKEMON_CATEGORY}/groups", timeout=30, headers=HEADERS)
    r.raise_for_status()
    d = r.json()
    return d.get("results", d) if isinstance(d, dict) else d

def fetch_tcgcsv_products(gid):
    r = requests.get(f"{TCGCSV_BASE}/{POKEMON_CATEGORY}/{gid}/products", timeout=60, headers=HEADERS)
    r.raise_for_status()
    d = r.json()
    return d.get("results", d) if isinstance(d, dict) else d

def fetch_tcgcsv_prices(gid):
    r = requests.get(f"{TCGCSV_BASE}/{POKEMON_CATEGORY}/{gid}/prices", timeout=60, headers=HEADERS)
    r.raise_for_status()
    d = r.json()
    return d.get("results", d) if isinstance(d, dict) else d

def get_ext(product, field_name):
    for item in product.get("extendedData", []):
        if item.get("name") == field_name:
            return item.get("value", "")
    return ""

def apply_prices(card_set, price_map, dry_run, stdout, style):
    stats = {"matched": 0, "updated": 0, "skipped": 0, "unmatched": 0}
    our_products = list(PokemonProduct.objects.filter(card_set=card_set))
    if not our_products:
        stdout.write(style.WARNING(f"  No products in DB for {card_set.code}"))
        return stats
    by_namekey = {name_key(p.name, p.card_number): p for p in our_products}
    updates = []
    for k, field_map in price_map.items():
        product = by_namekey.get(k)
        if not product:
            stats["unmatched"] += 1
            continue
        stats["matched"] += 1
        changed = False
        for field, zar in field_map.items():
            if zar != getattr(product, field, None):
                setattr(product, field, zar)
                changed = True
        # Always fix price=0 even if variant fields unchanged
        if not changed and (not product.price or product.price == 0):
            changed = True
        if changed:
            # Set price from variant fields
            if product.price_normal and not product.price_holo:
                product.price = product.price_normal
            elif product.price_holo and not product.price_normal:
                product.price = product.price_holo
            elif product.price_normal and product.price_holo:
                product.price = max(product.price_normal, product.price_holo)
            elif product.price_reverse_holo:
                product.price = product.price_reverse_holo
            elif product.price_first_edition:
                product.price = product.price_first_edition
            # Set the main price field based on ball variant
            vo = (product.variant_override or "").strip()
            if vo in ("BRH-PB", "RH-PB") and product.price_pokeball:
                product.price = product.price_pokeball
            elif vo in ("BRH-MB", "RH-MB") and product.price_masterball:
                product.price = product.price_masterball
            elif vo == "BRH-FB" and product.price_friendball:
                product.price = product.price_friendball
            elif vo == "BRH-LB" and product.price_loveball:
                product.price = product.price_loveball
            elif vo == "BRH-QB" and product.price_quickball:
                product.price = product.price_quickball
            elif vo == "BRH-DB" and product.price_duskball:
                product.price = product.price_duskball
            stats["updated"] += 1
            if not dry_run: updates.append(product)
        else:
            stats["skipped"] += 1
    if updates and not dry_run:
        with transaction.atomic():
            PokemonProduct.objects.bulk_update(updates, ALL_PRICE_FIELDS + ["price", "price_pokeball", "price_masterball", "price_friendball", "price_loveball", "price_quickball", "price_duskball"])
    return stats

def load_xlsx(xlsx_path):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    data = {}
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        era, set_name, abbrev, group_id, product_id, name, number, rarity, card_type, stage, variant, market_usd, low_usd, zar_val, rate, is_card = row
        if not abbrev or not name: continue
        if zar_val is None: skipped += 1; continue
        db_code = XLSX_TO_DB.get(abbrev, abbrev)
        field = variant_to_field(variant or "Normal")
        zar = Decimal(str(zar_val))
        k = name_key(name, number)
        if db_code not in data: data[db_code] = {}
        if k not in data[db_code]: data[db_code][k] = {}
        if zar > data[db_code][k].get(field, Decimal("0")):
            data[db_code][k][field] = zar
    wb.close()
    return data, skipped

def build_price_map_from_tcgcsv(tcg_products, tcg_prices, rate):
    pid_to_info = {}
    for p in tcg_products:
        num_raw = get_ext(p, "Number")
        pid_to_info[p["productId"]] = {"name": p.get("name", "").strip(), "number": norm_num(num_raw)}
    pid_to_usd = {}
    for row in tcg_prices:
        pid = row.get("productId")
        field = variant_to_field(row.get("subTypeName", "Normal"))
        usd_raw = row.get("marketPrice") or row.get("midPrice") or row.get("lowPrice")
        if not pid or not usd_raw: continue
        usd = Decimal(str(usd_raw))
        if usd <= 0: continue
        if pid not in pid_to_usd: pid_to_usd[pid] = {}
        if usd > pid_to_usd[pid].get(field, Decimal("0")): pid_to_usd[pid][field] = usd
    price_map = {}
    for pid, field_usd in pid_to_usd.items():
        info = pid_to_info.get(pid)
        if not info: continue
        k = name_key(info["name"], info["number"])
        if k not in price_map: price_map[k] = {}
        for field, usd in field_usd.items():
            zar = to_zar(usd, rate)
            if zar > price_map[k].get(field, Decimal("0")): price_map[k][field] = zar
    return price_map

def resolve_gid(cs, by_name, by_abbr):
    n = (cs.name or "").strip().lower()
    c = (cs.code or "").strip().upper()
    if n in by_name: return by_name[n]
    if c in by_abbr: return by_abbr[c]
    for tn, gid in by_name.items():
        if n in tn or tn in n: return gid
    return None

class Command(BaseCommand):
    help = "Sync prices from xlsx or live TCGCSV. ZAR = USD x rate x 1.1 up to R0.50."

    def add_arguments(self, parser):
        parser.add_argument("--set-code", dest="set_code", default=None)
        parser.add_argument("--xlsx", dest="xlsx", default=None)
        parser.add_argument("--rate", dest="rate", type=float, default=None)
        parser.add_argument("--dry-run", dest="dry_run", action="store_true", default=False)
        parser.add_argument("--delay", dest="delay", type=float, default=0.25)

    def handle(self, *args, **options):
        dry_run = options["dry_run"]
        if dry_run: self.stdout.write(self.style.WARNING("DRY RUN — no writes"))
        if options["xlsx"]:
            self._handle_xlsx(options["xlsx"], options["set_code"], dry_run)
        else:
            self._handle_live(options["set_code"], options["rate"], dry_run, options["delay"])

    def _handle_xlsx(self, xlsx_path, set_code, dry_run):
        if not os.path.exists(xlsx_path):
            raise CommandError(f"xlsx not found: {xlsx_path}")
        self.stdout.write(f"Loading xlsx: {xlsx_path}")
        xlsx_data, skipped = load_xlsx(xlsx_path)
        self.stdout.write(f"  {sum(len(v) for v in xlsx_data.values())} cards across {len(xlsx_data)} sets ({skipped} rows skipped).")
        if set_code:
            code = set_code.upper()
            if code not in xlsx_data:
                raise CommandError(f"'{code}' not in xlsx. Available: {sorted(xlsx_data.keys())}")
            sets_to_sync = {code: xlsx_data[code]}
        else:
            sets_to_sync = xlsx_data
        totals = {"matched": 0, "updated": 0, "skipped": 0, "unmatched": 0}
        for db_code, price_map in sets_to_sync.items():
            try:
                cs = CardSet.objects.get(code=db_code)
            except CardSet.DoesNotExist:
                self.stdout.write(self.style.WARNING(f"[{db_code}] not in DB — skipping"))
                continue
            self.stdout.write(f"[{db_code}] {len(price_map)} cards in xlsx")
            s = apply_prices(cs, price_map, dry_run, self.stdout, self.style)
            for k in totals: totals[k] += s[k]
            self.stdout.write(self.style.SUCCESS(f"  matched={s['matched']} updated={s['updated']} skipped={s['skipped']} unmatched={s['unmatched']}"))
        self._totals(totals, dry_run)

    def _handle_live(self, set_code, rate_override, dry_run, delay):
        rate = Decimal(str(rate_override)) if rate_override else fetch_rate()
        self.stdout.write(self.style.SUCCESS(f"1 USD = R{rate}"))
        self.stdout.write("Fetching TCGCSV groups...")
        try:
            groups = fetch_groups()
        except Exception as e:
            raise CommandError(f"Failed: {e}")
        self.stdout.write(f"  {len(groups)} groups found.")
        by_name = {(g.get("name") or "").strip().lower(): g.get("groupId") or g.get("id") for g in groups}
        by_abbr = {(g.get("abbreviation") or "").strip().upper(): g.get("groupId") or g.get("id") for g in groups}
        sets = [CardSet.objects.get(code=set_code.upper())] if set_code else list(CardSet.objects.all().order_by("release_date"))
        totals = {"matched": 0, "updated": 0, "skipped": 0, "unmatched": 0}
        for cs in sets:
            gid = resolve_gid(cs, by_name, by_abbr)
            if not gid:
                self.stdout.write(self.style.WARNING(f"[{cs.code}] no TCGCSV match — skipping"))
                continue
            self.stdout.write(f"[{cs.code}] group={gid}")
            try:
                tcg_prods = fetch_tcgcsv_products(gid)
                time.sleep(delay)
                tcg_prices = fetch_tcgcsv_prices(gid)
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"  {e} — skipping"))
                time.sleep(delay)
                continue
            self.stdout.write(f"  {len(tcg_prods)} products, {len(tcg_prices)} price rows")
            price_map = build_price_map_from_tcgcsv(tcg_prods, tcg_prices, rate)
            s = apply_prices(cs, price_map, dry_run, self.stdout, self.style)
            for k in totals: totals[k] += s[k]
            self.stdout.write(self.style.SUCCESS(f"  matched={s['matched']} updated={s['updated']} skipped={s['skipped']} unmatched={s['unmatched']}"))
            time.sleep(delay)
        self._totals(totals, dry_run)

    def _totals(self, totals, dry_run):
        self.stdout.write(self.style.SUCCESS(f"\nDONE  matched={totals['matched']}  updated={totals['updated']}  skipped={totals['skipped']}  unmatched={totals['unmatched']}"))
        if dry_run: self.stdout.write(self.style.WARNING("Dry run — nothing written."))






